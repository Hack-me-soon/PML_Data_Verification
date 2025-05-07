from contextlib import contextmanager
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse
from app.logic import *
from app.models import SheetSelection
from fastapi import APIRouter
from fastapi.middleware.cors import CORSMiddleware
import json
import os
import sys
import json
import traceback
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from pydantic import BaseModel
from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from contextlib import redirect_stdout, redirect_stderr



class SheetSelection(BaseModel):
    current_sheet: str
    previous_sheet: str
    increment_sheet: str
    master_sheet: str

app = FastAPI()
UPLOAD_DIR = "uploaded"

LOG_FILE_PATH = Path("logs") / "process_operation.log"
LOG_FILE_PATH.parent.mkdir(exist_ok=True)

# Allow frontend JS to call backend APIs (if needed)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Replace with specific origin in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# This directory points to the app/ folder
BASE_DIR = os.path.dirname(__file__)

# Mount static files (CSS, JS, etc.)
app.mount("/static", StaticFiles(directory=BASE_DIR), name="static")

# Point Jinja2 to app/ folder (where index.html is)
templates = Jinja2Templates(directory=BASE_DIR)

# Serve index.html at root
@app.get("/", response_class=HTMLResponse)
async def get_index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

os.makedirs("uploaded", exist_ok=True)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # or ["*"] for dev
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/upload-files/")
async def upload_files(
    master_data: UploadFile = File(...),
    increment_data: UploadFile = File(...),
    current_month_data: UploadFile = File(...),
    previous_month_data: UploadFile = File(...)
):
    clear_uploaded_folder(UPLOAD_DIR)

    uploads = {
        "master_file": master_data,
        "increment_file": increment_data,
        "current_file": current_month_data,
        "previous_file": previous_month_data
    }

    saved_paths = {}

    for label, file in uploads.items():
        filename = f"{label}_{file.filename}"
        full_path = os.path.join("uploaded", filename)

        with open(full_path, "wb") as f:
            f.write(await file.read())

        converted_path = convert_if_needed(full_path)
        saved_paths[label] = os.path.basename(converted_path)  # Store only filename

    return extract_sheet_names({k: os.path.join("uploaded", v) for k, v in saved_paths.items()})


@app.post("/process-selected-sheets/")
async def process_selected(selection: SheetSelection = None):
    base_dir = "uploaded"
    selection_path = os.path.join(base_dir, "selection.json")

    if selection:
        # Get the latest matching files from uploaded folder
        file_map = {
            "current_file": max([f for f in os.listdir(base_dir) if f.startswith("current_file_")], default=None),
            "previous_file": max([f for f in os.listdir(base_dir) if f.startswith("previous_file_")], default=None),
            "increment_file": max([f for f in os.listdir(base_dir) if f.startswith("increment_file_")], default=None),
            "master_file": max([f for f in os.listdir(base_dir) if f.startswith("master_file_")], default=None),
        }

        # Check if any required files are missing
        if not all(file_map.values()):
            raise FileNotFoundError("Some required uploaded files are missing.")

        print(f"File map: {file_map}")  # Debugging: Print the file map to verify which file is being selected

        # Combine both sheet names and actual filenames
        combined = {
            **file_map,
            "current_sheet": selection.current_sheet,
            "previous_sheet": selection.previous_sheet,
            "increment_sheet": selection.increment_sheet,
            "master_sheet": selection.master_sheet
        }

        # Save the selection file as JSON
        with open(selection_path, "w") as f:
            json.dump(combined, f)

        return {"message": "Sheets selected and saved.", "selection": combined}

    # If no selection, create auto-selection from the first sheet in each file
    auto_selection = ensure_selection_file(base_dir)
    return {
        "message": "selection.json was missing and created using the first sheet of each uploaded file.",
        "selection": auto_selection
    }


@app.post("/process-operation/")
async def process_operation():
    try:
        with open(LOG_FILE_PATH, "w", encoding="utf-8") as log_file:
            with redirect_stdout(log_file), redirect_stderr(log_file):
                print("[INFO] Process started.")
                # Load and prepare all data
                dfs = load_and_prepare_all_data()
                merged_df = merge_employee_data(dfs)

                # Export current_df to Excel
                output_path = "output.xlsx"
                current_df = global_context["current_df"]

                # Read selected sheet name from selection.json
                with open("uploaded/selection.json", "r") as f:
                    selection_data = json.load(f)
                current_sheet_name = selection_data.get("current_sheet")

                export_info = export_current_df_to_excel(current_df, current_sheet_name, output_path)

                # Load Excel sheet again for mismatch checks
                wb = load_workbook(output_path)
                ws_main = wb[current_sheet_name]

                result = export_current_df_to_excel(current_df, current_sheet_name, output_path)
                wb = result["wb"]                         # Use this instead of load_workbook()
                ws_error = result["ws_error"]
                ws_main = wb[current_sheet_name]         # This still works if `current_sheet_name` is valid

                # Extract headers
                headers = [str(cell.value).strip().lower().replace(" ", "") for cell in ws_main[1]]

                # Calculate ESI wages and detect mismatches
                esi_wages_list, total_wages_error_rows = calculate_esi_eligibility(
                    current_df, merged_df, ws_main, headers
                )

                # Processs Increment file with Added Increments
                increment_df = process_increment_file(dfs)

                # Calculate ESI wages
                calculate_esi_wages(
                    wb=wb,
                    ws_main=ws_main,
                    ws_error=ws_error,
                    esi_wages_list=esi_wages_list,
                    previous_df=dfs["previous_file"],
                    master_df=dfs["master_file"],
                    increment_df=increment_df,
                    output_file=output_path
                )

                calculate_pf_wages(
                    wb=wb,
                    ws_main=ws_main,
                    ws_error=ws_error,
                    previous_df=dfs["previous_file"],
                    master_df=dfs["master_file"],
                    increment_df=increment_df,
                    headers=headers,
                    merged_df=merged_df
                )

                wb.save(output_path)
                print(f"✅ Final report saved as '{output_path}'")

        return get_output_excel_response()

    except Exception as e:
        error_msg = f"[ERROR] {e}\n{traceback.format_exc()}"
        with open(LOG_FILE_PATH, "a", encoding="utf-8") as log_file:
            log_file.write(error_msg)
        return JSONResponse(content={"error": str(e)}, status_code=500)
    
@app.get("/get-latest-log/")
async def get_latest_log():
    if LOG_FILE_PATH.exists():
        with open(LOG_FILE_PATH, "r", encoding="utf-8") as f:
            return {"log": f.read()}
    return {"log": "[No logs found]"}


router = APIRouter()
