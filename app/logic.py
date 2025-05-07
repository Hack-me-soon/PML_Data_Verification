# app/logic.py

import os
import pandas as pd
from typing import Dict
import pandas as pd
import re
import openpyxl
import json
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
from math import ceil
from fastapi.responses import StreamingResponse


# Function to delete all files in the uploaded directory
def clear_uploaded_folder(UPLOAD_DIR):
    # Ensure the folder exists
    if os.path.exists(UPLOAD_DIR):
        for filename in os.listdir(UPLOAD_DIR):
            file_path = os.path.join(UPLOAD_DIR, filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path)  # Delete the file
                elif os.path.isdir(file_path):
                    os.rmdir(file_path)  # Remove directories if any (although there shouldn't be any)
            except Exception as e:
                print(f"Error deleting {file_path}: {e}")
    else:
        os.makedirs(UPLOAD_DIR)  # Create the folder if it doesn't exist


def convert_if_needed(file_path):
    if file_path.endswith(".xls"):
        xlsx_path = file_path.replace(".xls", ".xlsx")
        xls = pd.read_excel(file_path, sheet_name=None, engine='xlrd')  # Read all sheets
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            for sheet_name, df in xls.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        return xlsx_path
    return file_path

def extract_sheet_names(file_dict):
    sheet_names = {}
    for label, path in file_dict.items():
        try:
            xls = pd.ExcelFile(path)
            sheet_names[label] = xls.sheet_names
        except Exception as e:
            sheet_names[label] = [f"Error: {str(e)}"]
    return sheet_names

def load_excel_with_dynamic_header(file_path, sheet_name, required_headers):
    print(f"[INFO] Loading file: {file_path} | Sheet: {sheet_name}")
    temp_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=6, header=None)
    header_row_index = None

    for i in range(6):
        row = temp_df.iloc[i].astype(str).str.lower().str.strip()
        if all(h in row.values for h in required_headers):
            header_row_index = i
            print(f"[INFO] Header found in '{file_path}' (Sheet: {sheet_name}) at row {i + 1}")
            break

    if header_row_index is None:
        raise ValueError(f"[ERROR] Header with {required_headers} not found in first 6 rows of {file_path}")

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index)
    print(f"[INFO] Loaded {len(df)} rows from '{file_path}'\n")
    return df, header_row_index

def normalize_headers(df):
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "")
    return df

def get_uploaded_excel_files(base_dir="uploaded") -> dict:
    """
    Returns a dict of the expected uploaded Excel files from the base directory.
    """
    files = os.listdir(base_dir)
    matched_files = {
        "current_file": next((f for f in files if f.startswith("current_file_")), None),
        "previous_file": next((f for f in files if f.startswith("previous_file_")), None),
        "increment_file": next((f for f in files if f.startswith("increment_file_")), None),
        "master_file": next((f for f in files if f.startswith("master_file_")), None),
    }
    return matched_files


def load_and_prepare_all_data(base_dir="uploaded"):
    import json

    selection_path = os.path.join(base_dir, "selection.json")
    if not os.path.exists(selection_path):
        raise FileNotFoundError("selection.json not found in 'uploaded/'.")

    with open(selection_path, "r") as f:
        selection = json.load(f)

    current_sheet = selection["current_sheet"]
    previous_sheet = selection["previous_sheet"]
    increment_sheet = selection["increment_sheet"]
    master_sheet = selection["master_sheet"]

    matched_files = {
        key: selection[key]
        for key in ["current_file", "previous_file", "increment_file", "master_file"]
    }


    if not all(matched_files.values()):
        raise FileNotFoundError("One or more required Excel files are missing in 'uploaded/'.")

    # Define required headers for each file
    header_requirements = {
        "current_file": ["emp id", "rpd", "ot hour", "ot wages", "wrd days", "nfh",
                         "earned basic", "earned hra", "earned conveyance", "earned wa",
                         "earned other allowance", "total wages", "esi", "pf"],
        "previous_file": ["emp id", "rpd"],
        "increment_file": ["id", "inc. amt", "basic", "hra", "conv", "wa", "other allow"],
        "master_file": ["emp id", "rpd", "basicl1", "hral1", "conveyance", "wa", "other allowance", "grade"]
    }

    dfs = {}

    for key, filename in matched_files.items():
        file_path = os.path.join(base_dir, filename)
        sheet_name = selection[key.replace("_file", "_sheet")]
        required_headers = header_requirements[key]

        df, header_row = load_excel_with_dynamic_header(file_path, sheet_name, required_headers)
        df = normalize_headers(df)
        dfs[key] = df

    dfs["selected_sheets"] = selection  # Add this line to carry sheet names forward
    dfs["file_map"] = matched_files
    print("[SUCCESS] All files loaded and headers normalized.\n")
    return dfs


global_context = {}  # Optional: to hold DataFrames globally if needed later

def merge_employee_data(dfs):
    print("[INFO] Starting merge process...\n")

    current_df = dfs["current_file"]
    previous_df = dfs["previous_file"]
    master_df = dfs["master_file"]

    global_context["current_df"] = current_df
    global_context["previous_df"] = previous_df
    global_context["master_df"] = master_df

    # Merge current and previous data on empid
    print("[INFO] Merging current and previous wage data on 'empid'...\n")
    merged_df = pd.merge(current_df, previous_df, on='empid', how='left', suffixes=('_current', '_previous'))

    # Merge with master to get grade
    print("[INFO] Merging with master data to get 'grade'...\n")
    if 'empid' not in master_df.columns or 'grade' not in master_df.columns:
        raise KeyError("Required columns 'empid' or 'grade' not found in master file.")

    merged_df = pd.merge(merged_df, master_df[['empid', 'grade']], on='empid', how='left')

    # Ensure 'grade_current' is consistently named
    merged_df.rename(columns={'grade_y': 'grade_current'}, inplace=True)

    print(f"[SUCCESS] Merged DataFrame shape: {merged_df.shape}")

    global_context["merged_df"] = merged_df
    return merged_df


def export_current_df_to_excel(current_df, current_sheet_name, output_file):
    print(f"[INFO] Exporting current_df to Excel: {output_file}")

    # Write to Excel
    current_df.to_excel(output_file, index=False)

    # Load and modify workbook
    wb = load_workbook(output_file)
    ws_main = wb.active
    ws_main.title = current_sheet_name

    # === Create Error Sheet ===
    error_sheet_name = f"{current_sheet_name}Error"
    if error_sheet_name in wb.sheetnames:
        del wb[error_sheet_name]
    ws_error = wb.create_sheet(error_sheet_name)


    # Create font for future formatting
    red_font = Font(color="FF0000")

    # Read header row to locate column indices
    headers = [str(cell.value).strip().lower().replace(" ", "") for cell in ws_main[1]]

    try:
        id_col_index = headers.index('empid') + 1
        rpd_col_index = headers.index('rpd') + 1
        ot_wages_col_index = headers.index('otwages') + 1
    except ValueError as e:
        raise ValueError(f"[ERROR] Required column not found in headers: {e}")

    print(f"[INFO] Header positions - empid: {id_col_index}, rpd: {rpd_col_index}, otwages: {ot_wages_col_index}")

    wb.save(output_file)
    print(f"[SUCCESS] Excel saved as {output_file}\n")

    return {
        "wb": wb,
        "output_file": output_file,
        "sheet_name": current_sheet_name,
        "column_indices": {
            "empid": id_col_index,
            "rpd": rpd_col_index,
            "otwages": ot_wages_col_index
        },
        "ws_error": ws_error  # <-- Add this line to return ws_error
    }

def ensure_selection_file(base_dir="uploaded") -> dict:
    """
    Ensure 'selection.json' exists in the given base_dir.
    If not present, auto-create it using the first sheet from each relevant file.
    Returns the final selection dict.
    """
    selection_path = os.path.join(base_dir, "selection.json")
    
    if os.path.exists(selection_path):
        with open(selection_path, "r") as f:
            return json.load(f)

    # Auto-create selection from first sheet in each file
    matched_files = get_uploaded_excel_files(base_dir)

    if not all(matched_files.values()):
        raise FileNotFoundError("[ERROR] Required Excel files not found in 'uploaded/'.")

    def get_first_sheet(file_path):
        return pd.ExcelFile(file_path).sheet_names[0]

    auto_selection = {
        "current_sheet": get_first_sheet(os.path.join(base_dir, matched_files["current_file"])),
        "previous_sheet": get_first_sheet(os.path.join(base_dir, matched_files["previous_file"])),
        "increment_sheet": get_first_sheet(os.path.join(base_dir, matched_files["increment_file"])),
        "master_sheet": get_first_sheet(os.path.join(base_dir, matched_files["master_file"])),
    }

    with open(selection_path, "w") as f:
        json.dump(auto_selection, f, indent=4)

    print("[INFO] selection.json was created automatically using the first sheets from each file.")
    return auto_selection

def calculate_esi_eligibility(current_df, merged_df, ws_main, headers):

    print("\n[INFO] Starting Total ESI Wages Eligibility calculation...")
    esi_wages_list = []
    total_wages_error_rows = []

    # Identify column positions
    id_col_index = next((i + 1 for i, h in enumerate(headers) if h == 'empid'), None)
    total_wages_col_index = next((i + 1 for i, h in enumerate(headers) if h == 'totalwages'), None)

    if total_wages_col_index is None:
        print("[WARNING] Required column 'Total Wages' not found. Skipping ESI wage check.")
        return esi_wages_list, total_wages_error_rows

    for index, row in current_df.iterrows():
        try:
            emp_id = str(row['empid']).strip()
            rpd = float(row['rpd'])

            # Get grade from merged_df
            merged_row = merged_df[merged_df['empid'].astype(str).str.strip() == emp_id]
            grade = str(merged_row.iloc[0]['grade_current']).strip().lower() if not merged_row.empty else ""

            # ESI Wages Logic
            if grade in ("staff", "foreman", "graduate assistant"):
                esi_wages = rpd
                # print(f"[ESI] Emp ID {emp_id}: grade=staff → esi_wages = rpd = {rpd}")
            else:
                esi_wages = round(rpd * 26)
                # print(f"[ESI] Emp ID {emp_id}: grade=worker → esi_wages = round(rpd * 26) = {esi_wages}")

            esi_wages_list.append((emp_id, esi_wages))

            # Compare with Excel sheet value
            current_sheet_value = None
            for i in range(2, ws_main.max_row + 1):
                if str(ws_main.cell(row=i, column=id_col_index).value).strip() == emp_id:
                    current_sheet_value = ws_main.cell(row=i, column=total_wages_col_index).value
                    try:
                        current_sheet_value = float(current_sheet_value)
                    except:
                        current_sheet_value = None

                    if current_sheet_value != esi_wages:
                        mismatch_msg = f"[ESI Mismatch] Expected: {esi_wages}, Found: {current_sheet_value} (but shown in Total Wages column)"
                        total_wages_error_rows.append([emp_id, mismatch_msg])
                    break

        except Exception as e:
            print(f"[ERROR] Error processing ESI wages for Employee ID {emp_id}: {e}")
            total_wages_error_rows.append([emp_id, f"Error: {e}"])

    # print(f"[INFO] Finished ESI wage calculation. {len(total_wages_error_rows)} mismatches found.\n")
    return esi_wages_list, total_wages_error_rows

def normalize_column(col_name):
    return re.sub(r'[^a-z0-9]', '', col_name.lower().strip())

def map_columns(columns):
    col_map = {}
    for col in columns:
        norm = normalize_column(col)
        if norm in ['id', 'empid', 'employeeid', 'idno', 'idnumber']:
            col_map['id'] = col
        elif 'basic' in norm:
            col_map['basic'] = col
        elif 'hra' in norm:
            col_map['hra'] = col
        elif 'conv' in norm:
            col_map['conv'] = col
        elif norm == 'wa':
            col_map['wa'] = col
        elif 'otherallow' in norm or ('other' in norm and 'allow' in norm):
            col_map['otherallow'] = col
    return col_map

def find_header_row_and_map(df):
    for i in range(min(6, len(df))):
        row = df.iloc[i].astype(str).tolist()
        col_map = map_columns(row)
        if 'id' in col_map:
            return i, col_map
    return None, None

def process_increment_file(dfs):
    increment_file = dfs["file_map"]["increment_file"]
    increment_sheet = dfs["selected_sheets"]["increment_sheet"]

    # Suppress specific openpyxl warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        xl = pd.ExcelFile(f"uploaded/{increment_file}")

    all_sheets = xl.sheet_names

    print(f"\n📄 Scanning increment file: {increment_file}")
    print(f"🗂️ Sheets found: {', '.join(all_sheets)}")

    if increment_sheet not in all_sheets:
        raise ValueError(f"{increment_sheet} not found in {increment_file}")

    end_index = all_sheets.index(increment_sheet)
    target_sheets = all_sheets[:end_index + 1]

    cumulative_data = []

    for sheet in target_sheets:
        print(f"\n🔍 Processing sheet: '{sheet}'")
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df_raw = xl.parse(sheet, header=None)
        header_row, col_map = find_header_row_and_map(df_raw)

        if header_row is None:
            print(f"❌ Header not found in sheet: '{sheet}' — Skipping")
            continue
        else:
            print(f"✅ Header found in sheet: '{sheet}' at row {header_row + 1}")

        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df = xl.parse(sheet, header=header_row)
        df.columns = df.columns.str.strip()

        selected_cols = {
            key: col_map[key]
            for key in col_map
            if key in ['id', 'basic', 'hra', 'conv', 'wa', 'otherallow']
        }

        if 'id' not in selected_cols:
            print(f"⚠️ Skipping sheet: '{sheet}' — No 'id' column after mapping")
            continue

        temp_df = df[list(selected_cols.values())].rename(columns={v: k for k, v in selected_cols.items()})
        temp_df['id'] = temp_df['id'].astype(str).str.extract(r'(\d+)')[0]
        temp_df['id'] = pd.to_numeric(temp_df['id'], errors='coerce')
        temp_df = temp_df.dropna(subset=['id'])
        temp_df['id'] = temp_df['id'].astype(int)

        for col in ['basic', 'hra', 'conv', 'wa', 'otherallow']:
            if col in temp_df:
                temp_df[col] = pd.to_numeric(temp_df[col], errors='coerce').fillna(0)
            else:
                temp_df[col] = 0

        temp_df = temp_df[['id', 'basic', 'hra', 'conv', 'wa', 'otherallow']]
        cumulative_data.append(temp_df)

    if not cumulative_data:
        print("\n⚠️ No valid data found in any sheets.")
        return pd.DataFrame(columns=['id', 'basic', 'hra', 'conv', 'wa', 'otherallow'])

    inc_merged_df = pd.concat(cumulative_data, ignore_index=True)
    final_df = inc_merged_df.groupby('id', as_index=False).sum(numeric_only=True)

    final_df.to_csv("increment_cleaned_data.csv", index=False)
    print("\n✅ Final cleaned increment data saved to 'increment_cleaned_data.csv'")

    return final_df

def custom_round(val):
    if val - int(val) >= 0.5:
        return int(val) + 1
    else:
        return int(val)

def calculate_esi_wages(wb, ws_main, ws_error, esi_wages_list, previous_df, master_df, increment_df, output_file):

    print("\n[INFO] Starting Total ESI Wages Eligibility calculation...")

    increment_df['id'] = increment_df['id'].astype(str).str.strip()

    esi_error_logged = False
    esi_error_row = ws_error.max_row + 3

    esi_wages_dict = {str(int(float(emp_id))): wage for emp_id, wage in esi_wages_list}

    headers = [ws_main.cell(row=1, column=col).value.strip().lower() for col in range(1, ws_main.max_column + 1)]

    id_col_index = next((i + 1 for i, h in enumerate(headers) if h in ['empid', 'employeeid']), None)
    esi_col_index = next((i + 1 for i, h in enumerate(headers) if h == 'esi'), None)

    if id_col_index is None or esi_col_index is None:
        print("[ERROR] 'EmployeeID' or 'ESI' column not found.")
        pass
    else:
        esi_wages_dict = {str(int(float(emp_id))): wage for emp_id, wage in esi_wages_list}

        for row in range(2, ws_main.max_row + 1):
            try:
                raw_id = ws_main.cell(row=row, column=id_col_index).value
                emp_id = str(int(float(raw_id))) if raw_id is not None else ""
                esi_wages = esi_wages_dict.get(emp_id, 0)

                earned_basic = float(ws_main.cell(row=row, column=headers.index("earnedbasic") + 1).value or 0)
                earned_hra = float(ws_main.cell(row=row, column=headers.index("earnedhra") + 1).value or 0)
                earned_other = float(ws_main.cell(row=row, column=headers.index("earnedotherallowance") + 1).value or 0)
                ot_wages = float(ws_main.cell(row=row, column=headers.index("otwages") + 1).value or 0)
                wrd_days = float(ws_main.cell(row=row, column=headers.index("wrddays") + 1).value or 0)
                nfh = float(ws_main.cell(row=row, column=headers.index("nfh") + 1).value or 0)
                rpd_current = float(ws_main.cell(row=row, column=headers.index("rpd") + 1).value or 0)

                prev_row = previous_df[previous_df['empid'].astype(str).str.strip() == emp_id]
                rpd_prev = float(prev_row.iloc[0]['rpd']) if not prev_row.empty else None

                master_row = master_df[master_df['empid'].astype(str).str.strip() == emp_id]
                if not master_row.empty:
                    basicL1 = float(master_row.iloc[0]['basicl1'])
                    hral1 = float(master_row.iloc[0]['hral1'])
                    wa = float(master_row.iloc[0]['wa'])
                    conveyance = float(master_row.iloc[0]['conveyance'])
                    other_allow = float(master_row.iloc[0].get('otherallowance', 0) or 0)
                    grade = master_row.iloc[0].get("grade", "").lower()
                else:
                    basicL1 = hral1 = wa = conveyance = other_allow = 0
                    grade = ""

                inc_row = increment_df[increment_df['id'] == emp_id]
                if not inc_row.empty:
                    inc_basic = float(inc_row.iloc[0]['basic'])
                    inc_hra = float(inc_row.iloc[0]['hra'])
                    inc_conv = float(inc_row.iloc[0]['conv'])
                    inc_wa = float(inc_row.iloc[0]['wa'])
                    inc_other = float(inc_row.iloc[0]['otherallow'])
                else:
                    inc_basic = inc_hra = inc_conv = inc_wa = inc_other = 0

                if esi_wages >= 21000:
                    esi_calculated = 0
                else:
                    if grade in ("staff", "foreman", "graduate assistant"):
                        if pd.notna(rpd_prev) and rpd_current == rpd_prev:
                            esi_wages = earned_basic + earned_hra + earned_other + ot_wages
                            # print(f"[INFO]Staff Emp ID {emp_id}: Using Current Earned → "
                            #         f"earned_basic = {earned_basic}, earned_hra = {earned_hra}, "
                            #         f"earned_other = {earned_other}, ot_wages = {ot_wages}")
                        else:
                            esi_master_total = (basicL1 + hral1 + other_allow) / 26
                            esi_inc_total = (inc_basic + inc_hra + inc_other) / 26
                            esi_wages = ((esi_master_total + esi_inc_total) * (wrd_days + nfh)) + ot_wages
                            # print(f"[INFO]Staff Emp ID {emp_id}: Using Master + Increment → "
                            #         f"basicL1 = {basicL1}, hral1 = {hral1}, other_allow = {other_allow}, "
                            #         f"inc_basic = {inc_basic}, inc_hra = {inc_hra}, inc_other = {inc_other}, "
                            #         f"wrd_days = {wrd_days}, nfh = {nfh}, ot_wages = {ot_wages}")
                    else:
                        if pd.notna(rpd_prev) and rpd_current == rpd_prev:
                            esi_wages = earned_basic + earned_hra + earned_other + ot_wages
                            # print(f"[INFO] Emp ID {emp_id}: Using Current Earned → "
                            #         f"earned_basic = {earned_basic}, earned_hra = {earned_hra}, "
                            #         f"earned_other = {earned_other}, ot_wages = {ot_wages}")
                        else:
                            esi_wages = (
                                (basicL1 + hral1 + other_allow) *(wrd_days + nfh) +
                                (inc_basic + inc_hra + inc_other) * (wrd_days + nfh) +
                                ot_wages
                            )
                            # print(f"[INFO] Emp ID {emp_id}: Using Master + Increment → "
                            #         f"basicL1 = {basicL1}, hral1 = {hral1}, other_allow = {other_allow}, "
                            #         f"inc_basic = {inc_basic}, inc_hra = {inc_hra}, inc_other = {inc_other}, "
                            #         f"wrd_days = {wrd_days}, nfh = {nfh}, ot_wages = {ot_wages}")

                    esi_calculated = ceil(esi_wages * 0.0075)
                    # print("esi =",esi_calculated)

                # Compare with recorded ESI
                matched_row = None
                esi_recorded = None
                for r in range(2, ws_main.max_row + 1):
                    if str(ws_main.cell(row=r, column=id_col_index).value).strip() == emp_id:
                        esi_recorded = ws_main.cell(row=r, column=esi_col_index).value
                        matched_row = r
                        try:
                            esi_recorded = float(esi_recorded)
                        except:
                            esi_recorded = None
                        break

                if esi_calculated != esi_recorded and matched_row:
                    # print(f"[LOG] Mismatch found for Emp ID: {emp_id} | Recorded: {esi_recorded}, Calculated: {esi_calculated}")

                    ws_main.cell(row=matched_row, column=esi_col_index).font = Font(color="FF0000", bold=True)

                    if not esi_error_logged:
                        # print("[LOG] Writing ESI error headers to error sheet...")
                        ws_error.cell(row=esi_error_row, column=1).value = "ESI Error:"
                        ws_error.cell(row=esi_error_row + 1, column=1).value = "Emp ID"
                        ws_error.cell(row=esi_error_row + 1, column=2).value = "Given ESI"
                        ws_error.cell(row=esi_error_row + 1, column=3).value = "Calculated ESI"
                        esi_error_row += 2
                        esi_error_logged = True

                    emp_name = ws_main.cell(row=matched_row, column=headers.index("name") + 1).value if "name" in headers else ""
                    # print(f"[LOG] Logging mismatch for: {emp_id} ({emp_name})")

                    ws_error.cell(row=esi_error_row, column=1).value = emp_id
                    ws_error.cell(row=esi_error_row, column=2).value = esi_recorded
                    ws_error.cell(row=esi_error_row, column=3).value = esi_calculated
                    esi_error_row += 1

                
            except Exception as e:
                print(f"[ERROR] Error processing Employee ID {emp_id}: {e}")

def calculate_pf_wages(wb, ws_main, ws_error, previous_df, master_df, increment_df, headers, merged_df):
    print("\n[INFO] Starting PF Wages Validation...")

    increment_df['id'] = increment_df['id'].astype(str).str.strip()

    pf_error_logged = False
    pf_error_row = ws_error.max_row + 3

    id_col_index = next((i + 1 for i, h in enumerate(headers) if h in ['empid', 'employeeid']), None)
    pf_col_index = headers.index("pf") + 1 if "pf" in headers else None

    if id_col_index is None or pf_col_index is None:
        print("[ERROR] 'EmployeeID' or 'PF' column not found.")
    else:
        for row in range(2, ws_main.max_row + 1):
            try:
                raw_id = ws_main.cell(row=row, column=id_col_index).value
                emp_id = str(raw_id).strip() if raw_id is not None else ""

                earned_basic = float(ws_main.cell(row=row, column=headers.index("earnedbasic") + 1).value or 0)
                earned_conv = float(ws_main.cell(row=row, column=headers.index("earnedconveyance") + 1).value or 0)
                earned_other = float(ws_main.cell(row=row, column=headers.index("earnedotherallowance") + 1).value or 0)
                wrd_days = float(ws_main.cell(row=row, column=headers.index("wrddays") + 1).value or 0)
                nfh = float(ws_main.cell(row=row, column=headers.index("nfh") + 1).value or 0)
                rpd_current = float(ws_main.cell(row=row, column=headers.index("rpd") + 1).value or 0)

                prev_row = previous_df[previous_df['empid'].astype(str).str.strip() == emp_id]
                rpd_prev = float(prev_row.iloc[0]['rpd']) if not prev_row.empty else None

                master_row = master_df[master_df['empid'].astype(str).str.strip() == emp_id]
                basicL1 = float(master_row.iloc[0]['basicl1']) if not master_row.empty else 0
                conv_master = float(master_row.iloc[0]['conveyance']) if not master_row.empty else 0
                other_master = float(master_row.iloc[0]['otherallowance']) if not master_row.empty else 0

                inc_row = increment_df[increment_df['id'] == emp_id]
                inc_basic = float(inc_row.iloc[0]['basic']) if not inc_row.empty else 0
                inc_conv = float(inc_row.iloc[0]['conv']) if not inc_row.empty else 0
                inc_other = float(inc_row.iloc[0]['otherallow']) if not inc_row.empty else 0

                merged_row = merged_df[merged_df['empid'].astype(str).str.strip() == emp_id]
                grade = str(merged_row.iloc[0]['grade_current']).strip().lower() if not merged_row.empty else ""

                # ==== PF Earned Basic Calculation ====
                if grade in ("staff", "foreman", "graduate assistant"):
                    if pd.notna(rpd_prev) and rpd_current == rpd_prev:
                        staff_earned_basic = earned_basic + earned_conv + earned_other
                        # print(f"[PF][STAFF] Emp ID {emp_id}: Same RPD → staff_earned_basic = earned_basic + conv + other = {staff_earned_basic}")
                    else:
                        master_total = (basicL1 + conv_master + other_master) / 26
                        inc_total = (inc_basic + inc_conv + inc_other) / 26
                        staff_earned_basic = (master_total + inc_total)*(wrd_days + nfh)
                        # print(f"master total={master_total}, increment total={inc_total}, [PF][STAFF] Emp ID {emp_id}: Diff RPD → staff_earned_basic = master_total + inc_total * wrd_days+nfh = "
                        #       f"{master_total} + {inc_total} * {wrd_days + nfh} = {staff_earned_basic}")

                    if staff_earned_basic > 15000:
                        staff_earned_basic = 15000
                        # print(f"[PF][STAFF] Emp ID {emp_id}: Capped to 15000")

                    pf_calculated = custom_round(staff_earned_basic * 0.12)
                    # print(f"[PF][STAFF] Emp ID {emp_id}: PF = custom_round({staff_earned_basic} * 0.12) = {pf_calculated}")

                else:
                    if pd.notna(rpd_prev) and rpd_current == rpd_prev:
                        pf_earned_basic = earned_basic
                        # print(f"[PF][WORKER] Emp ID {emp_id}: Same RPD → pf_earned_basic = earned_basic = {earned_basic}")
                    else:
                        pf_earned_basic = (basicL1 + inc_basic) * (wrd_days + nfh)
                        # print(f"[PF][WORKER] Emp ID {emp_id}: Diff RPD → pf_earned_basic = (basicL1 + inc_basic) * (wrd_days + nfh) = "
                        #       f"({basicL1} + {inc_basic}) * ({wrd_days} + {nfh}) = {pf_earned_basic}")

                    if pf_earned_basic > 15000:
                        pf_earned_basic = 15000
                        # print(f"[PF][WORKER] Emp ID {emp_id}: Capped to 15000")

                    pf_calculated = custom_round(pf_earned_basic * 0.12)
                    # print(f"[PF][WORKER] Emp ID {emp_id}: PF = custom_round({pf_earned_basic} * 0.12) = {pf_calculated}")

                # ===== Compare with existing PF column value =====
                pf_recorded = None
                if str(ws_main.cell(row=row, column=id_col_index).value).strip() == emp_id:
                    pf_recorded = ws_main.cell(row=row, column=pf_col_index).value
                    try:
                        pf_recorded = float(pf_recorded)
                    except:
                        pf_recorded = None

                if pf_calculated != pf_recorded:
                    pf_cell = ws_main.cell(row=row, column=pf_col_index)
                    pf_cell.font = Font(color="FF0000", bold=True)

                    if not pf_error_logged:
                        ws_error.cell(row=pf_error_row, column=1).value = "PF Error:"
                        ws_error.cell(row=pf_error_row + 1, column=1).value = "Emp ID"
                        ws_error.cell(row=pf_error_row + 1, column=2).value = "Given PF"
                        ws_error.cell(row=pf_error_row + 1, column=3).value = "Calculated PF"
                        pf_error_row += 2
                        pf_error_logged = True

                    emp_name = ws_main.cell(row=row, column=headers.index("name") + 1).value if "name" in headers else ""

                    ws_error.cell(row=pf_error_row, column=1).value = emp_id
                    ws_error.cell(row=pf_error_row, column=2).value = pf_recorded
                    ws_error.cell(row=pf_error_row, column=3).value = pf_calculated
                    pf_error_row += 1

            except Exception as e:
                print(f"[ERROR] Error processing PF for Employee ID {emp_id}: {e}")

def get_output_excel_response(file_path: str = "output.xlsx"):
    """
    Returns a StreamingResponse for the generated output.xlsx file.
    
    Parameters:
    - file_path: Path to the output Excel file (default: 'output.xlsx')
    
    Returns:
    - StreamingResponse: Excel file stream ready for FastAPI return
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_path} not found in root directory.")

    file_stream = open(file_path, "rb")
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={os.path.basename(file_path)}"}
    )